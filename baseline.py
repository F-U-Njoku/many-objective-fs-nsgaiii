import random
import pandas as pd
import openml as oml
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import cross_val_score
from sklearn.metrics import make_scorer, f1_score, precision_score, recall_score


def fetch_data(number):
    dataset = oml.datasets.get_dataset(number)
    X, y, categorical_indicator, attribute_names = dataset.get_data(
        dataset_format="array",
        target=dataset.default_target_attribute)
    return X, y, attribute_names, len(attribute_names)


def get_metric(classifier, X, y):
    if y.nunique() == 2:
        f1 = cross_val_score(estimator=classifier, X=X, y=y,
                             scoring="f1", cv=10).mean()
        precision = cross_val_score(estimator=classifier, X=X, y=y,
                                    scoring="precision", cv=10).mean()
        recall = cross_val_score(estimator=classifier, X=X, y=y,
                                 scoring="recall", cv=10).mean()
    else:
        f1 = cross_val_score(estimator=classifier, X=X, y=y,
                             scoring=make_scorer(f1_score, average='macro'), cv=10).mean()
        precision = cross_val_score(estimator=classifier, X=X, y=y, scoring=make_scorer(precision_score,
                                                                                        average='macro'), cv=10).mean()
        recall = cross_val_score(estimator=classifier, X=X, y=y, scoring=make_scorer(recall_score,
                                                                                     average='macro'), cv=10).mean()
    return precision, recall, f1


def experiment(data_id, classifier):
    X, y, attribute_names, feat = fetch_data(data_id)
    try:
        X = pd.DataFrame(X, columns=attribute_names)
    except:
        X = pd.DataFrame.sparse.from_spmatrix(X)
    y = pd.Series(y)
    X = X.dropna()
    if data_id == 41142:
        X.drop(columns=X.columns[X.nunique() < 100], inplace=True)
        attribute_names = X.columns

    # Select 3000 instances
    if data_id == 273:
        random.seed(1111)
        x_idx = random.sample(range(X.shape[0]), 1334)
        X = X.iloc[x_idx, :]
        attribute_names = X.columns
        y = y[x_idx]
    metrics = get_metric(classifier, X, y)
    classifier.fit(X, y)
    rank = pd.DataFrame(list(zip(attribute_names, classifier.feature_importances_)), columns=["features", "score"])
    rank.sort_values(by=['score'], ascending=False, inplace=True)

    wb2 = load_workbook(filename='results/master.xlsx')
    sheet2 = wb2["baseline"]
    sheet2.append([str(data_id), X.shape[1], X.shape[0], metrics[0], metrics[1], metrics[2]])
    wb2.save(filename='results/master.xlsx')

    with pd.ExcelWriter('results/master.xlsx', mode="a") as writer:
        rank.to_excel(writer, sheet_name=str(data_id), index=False)

    return


if __name__ == '__main__':
    wb = load_workbook(filename='results/master.xlsx')
    sheet = wb["baseline"]
    sheet.append(["data_id", "features", "instances", "precision", "recall", "f1"])
    wb.save(filename='results/master.xlsx')
    rf = RandomForestClassifier(max_features=None, random_state=0)
    for data in [1526, 40496, 1565, 1560, 1021, 995, 1004, 40666]:
        print("Data ID: ", data)
        experiment(data, rf)
    # wb.save(filename='results/master.xlsx')
